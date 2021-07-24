'''
Blank Row Inserter
Create a program blankRowInserter.py that takes two integers and a filename 
string as command line arguments. Let’s call the first integer N and the 
second integer M. Starting at row N, the program should insert M blank rows 
into the spreadsheet. For example, when the program is run like this:


python blankRowInserter.py 3 2 myProduce.xlsx
... the “before” and “after” spreadsheets should look like Figure 12-12.
'''
import openpyxl
import sys

#python blankRowInserter.py 3 2 myProduce.xlsx

m = int(sys.argv[1]) #posição
n = int(sys.argv[2]) #quantidade de linhas em branco
file_in = str(sys.argv[3])

wb_in = openpyxl.load_workbook(file_in)
sheet_in = wb_in.get_sheet_by_name('Sheet')

wb_out = openpyxl.Workbook()
sheet_out = wb_out.get_sheet_by_name('Sheet')

rows = 0
cols = 0

#contar a quantidade de linhas e colunas validas
for row in sheet_in.iter_rows():
	if cols == 0:
		cols = len(row)
	rows += 1

#copiar para novo arquivo de forma condicional de acordo com m e n
for i in range(1, rows + 1):
	for j in range (1, cols + 1):
		if i < m:
			sheet_out.cell(row=i, column=j).value = sheet_in.cell(row=i, column=j).value
		else:
			sheet_out.cell(row=i+n, column=j).value = sheet_in.cell(row=i, column=j).value

wb_out.save('New-' + file_in)