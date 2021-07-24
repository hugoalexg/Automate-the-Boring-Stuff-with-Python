'''
Multiplication Table Maker
Create a program multiplicationTable.py that takes a number N from the 
command line and creates an N×N multiplication table in an Excel spreadsheet. 
For example, when the program is run like this:

py multiplicationTable.py 6
... it should create a spreadsheet that looks like Figure 12-11.
'''
import sys
import openpyxl
from openpyxl.styles import Font

number = int(sys.argv[1])

wb = openpyxl.Workbook()

sheet = wb.get_sheet_by_name('Sheet')

for i in range(1, number + 1):
	sheet.cell(row=i+1, column=1).value = i
	sheet.cell(row=i+1, column=1).font = Font(bold=True)
	sheet.cell(row=1, column=i+1).value = i
	sheet.cell(row=1, column=i+1).font = Font(bold=True)

for i in range (1, number + 1):
	for j in range (1, number + 1):
		sheet.cell(row=i+1, column=j+1).value = sheet.cell(row=i+1, column=1).value * sheet.cell(row=1, column=j+1).value

wb.save('example2.xlsx')